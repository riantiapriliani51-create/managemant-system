from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session, flash, send_file
import os
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash

# ================= DATABASE =================
from db import db, Barang, User, LogTransaksi, init_db
from datetime import datetime

# Get the base directory of the application
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__, instance_path=os.path.join(BASE_DIR, 'instance'))
app.secret_key = 'inventori-login'

# Ensure instance folder exists
os.makedirs(app.instance_path, exist_ok=True)

# Database configuration - using absolute path for Windows compatibility
db_path = os.path.join(app.instance_path, 'inventory.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

init_db(app)

# ================= LOGOUT =================
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('dashboard'))

# ================= KEMBALI KE DASHBOARD =================
@app.route('/kembali-dashboard')
def kembali_dashboard():
    if 'role' not in session:
        return redirect(url_for('dashboard'))
    
    if session['role'] == 'admin':
        return redirect(url_for('dashboard_admin'))
    else:
        return redirect(url_for('dashboard_user'))

# ================= DASHBOARD (GAMBAR KE-2) =================
@app.route('/', methods=['GET', 'POST'])
def dashboard():
    if request.method == 'POST':
        user = User.query.filter_by(
            username=request.form['username']
        ).first()

        if user and check_password_hash(user.password, request.form['password']):
            session['user_id'] = user.id
            session['role'] = user.role
            session['username'] = user.username

            # ⬇️ INI YANG NAMPILKAN GAMBAR KEDUA
            if user.role == 'admin':
                return redirect(url_for('dashboard_admin'))
            else:
                return redirect(url_for('dashboard_user'))

        flash('Username atau password salah')

    # Tampilkan daftar ruang inventori di halaman login/dashboard (hanya lihat)
    # Selalu tampilkan ruang default meskipun database belum punya data untuk ruang tersebut.
    default_ruang = ['A', 'B', 'C', 'D']
    try:
        ruang_from_db = sorted({b.ruang for b in Barang.query.all() if b.ruang})
    except Exception:
        ruang_from_db = []

    # Gabungkan default dengan ruang dari DB tanpa membuat entri baru
    ruang_list = sorted(set(default_ruang + ruang_from_db))
    return render_template('dashboard.html', ruang_list=ruang_list)

# ================= ADMIN =================
@app.route('/admin')
def dashboard_admin():
    if 'role' not in session:
        return redirect(url_for('dashboard'))

    if session['role'] != 'admin':
        return redirect(url_for('dashboard'))

    return render_template('dashboard_admin.html')

# ================= USER =================
@app.route('/user')
def dashboard_user():
    if 'role' not in session:
        return redirect(url_for('dashboard'))

    if session['role'] != 'user':
        return redirect(url_for('dashboard'))

    return render_template('dashboard_user.html')

# ================= EXPORT EXCEL (HTML) =================
@app.route('/export')
def export_page():
    if session.get('role') not in ['admin', 'user']:
        return redirect(url_for('dashboard'))
    return render_template('export_excel.html')


# ================= EXPORT EXCEL (PROSES) =================
@app.route('/export-excel')
def export_excel():
    if session.get('role') not in ['admin', 'user']:
        return redirect(url_for('dashboard'))
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    import io
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventori"

    # ================= HEADER INVENTORI =================
    ws['B1'] = 'Nama'
    ws['C1'] = 'Jumlah'
    ws['D1'] = 'Unit'
    ws['E1'] = 'Ruang'

    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}1'].alignment = Alignment(horizontal='center')

    # ================= DATA INVENTORI =================
    row = 2
    items = Barang.query.order_by(Barang.id).all()

    for item in items:
        ws[f'B{row}'] = item.nama
        ws[f'C{row}'] = item.jumlah
        ws[f'D{row}'] = item.unit
        ws[f'E{row}'] = item.ruang

        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

        row += 1

    # ================= HEADER LOG =================
    log_start = row + 3

    ws[f'B{log_start}'] = 'Pengguna'
    ws[f'C{log_start}'] = 'Tanggal'
    ws[f'D{log_start}'] = 'Jam'
    ws[f'E{log_start}'] = 'Aksi'
    ws[f'F{log_start}'] = 'Jumlah'
    ws[f'G{log_start}'] = 'Unit'
    ws[f'H{log_start}'] = 'Jumlah Saat Ini'

    for col in ['B','C','D','E','F','G','H']:
        ws[f'{col}{log_start}'].alignment = Alignment(horizontal='center')

    # ================= DATA LOG =================
    logs = LogTransaksi.query.order_by(LogTransaksi.id).all()
    log_row = log_start + 1

    for log in logs:
        ws[f'B{log_row}'] = log.pengguna
        ws[f'C{log_row}'] = log.tanggal
        ws[f'D{log_row}'] = log.jam
        ws[f'E{log_row}'] = log.aksi
        ws[f'F{log_row}'] = log.jumlah
        ws[f'G{log_row}'] = log.unit
        ws[f'H{log_row}'] = log.jumlah_saat_ini

        for col in ['B','C','D','E','F','G','H']:
            ws[f'{col}{log_row}'].alignment = Alignment(horizontal='center')

        log_row += 1

    # ================= AUTO WIDTH =================
    for col in range(2, 9):  # kolom B - H
        col_letter = get_column_letter(col)
        max_length = 0

        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    # ================= EXPORT =================
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="inventori.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================= EXPORT EXCEL PER RUANG =================
@app.route('/export-inventori/<nama>')
def export_inventori(nama):
    # Hanya izinkan pengunjung (belum login) untuk mengakses export ruang A
    if session.get('role') is not None or nama.upper() != 'A':
        return redirect(url_for('dashboard'))
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    import io
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = f"Inventori {nama.upper()}"

    # ================= HEADER =================
    ws['B2'] = f'INVENTORI RUANG {nama.upper()}'
    ws['B2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('B2:F2')
    
    ws['B3'] = 'Nama'
    ws['C3'] = 'Jumlah'
    ws['D3'] = 'Unit'
    ws['E3'] = 'Min'
    ws['F3'] = 'Max'

    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].alignment = Alignment(horizontal='center')

    # ================= DATA =================
    row = 4
    items = Barang.query.filter_by(ruang=nama.upper()).order_by(Barang.id).all()

    for item in items:
        ws[f'B{row}'] = item.nama
        ws[f'C{row}'] = item.jumlah
        ws[f'D{row}'] = item.unit
        ws[f'E{row}'] = item.min_qty
        ws[f'F{row}'] = item.max_qty

        for col in ['C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        for col in ['B']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='left')

        row += 1

    # ================= AUTO WIDTH =================
    for col in range(1, 7):  # kolom A - F
        col_letter = get_column_letter(col)
        max_length = 0

        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    # ================= EXPORT =================
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"inventori_{nama.upper()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================= UPLOAD FOLDER =================
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif'}

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

# ================= INVENTORY B DENGAN UPLOAD PDF =================
@app.route('/inventori_b')
def inventori_b():
    # Baca file PDF yang sudah ada di folder uploads
    import glob
    pdf_files = glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], 'b_*.pdf'))
    pdf_list = [{'nama': os.path.basename(f).replace('b_', '').replace('.pdf', ''), 'file': os.path.basename(f)} for f in pdf_files]
    
    items_html = ''.join([
        f'''
        <div class="item-card">
            <div class="item-left">
                <h3>{item['nama']}</h3>
            </div>

            <div class="item-right">
                <div class="item-section">
                    <a href="/download/{item['file']}" class="btn-download">⬇ Download PDF</a>
                </div>
            </div>
        </div>
        '''
        for item in pdf_list
    ])

    return f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
    
        <meta charset="UTF-8">
        <title>Inventori B</title>
        <style>
            @keyframes moveText {{
                0% {{
                    transform: translateX(100%);
                }}
                100% {{
                    transform: translateX(-100%);
                }}
            }}

            body {{
                margin: 0;
                font-family: Arial, Helvetica, sans-serif;
                background-color: #f4f4f4;
            }}

            .header {{
                background: linear-gradient(90deg, #0b5fa5, #1e88e5);
                color: white;
                padding: 40px;
                text-align: center;
            }}

            .container {{
                padding: 40px;
                display: flex;
                gap: 30px;
            }}

            .list {{
                flex: 1;
            }}

            .sidebar {{
                width: 300px;
            }}

            .upload-box {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 3px 8px rgba(0,0,0,0.15);
                margin-bottom: 20px;
            }}

            .upload-box h3 {{
                margin-top: 0;
                color: #0b5fa5;
            }}

            .upload-box input,
            .upload-box button {{
                width: 100%;
                padding: 10px;
                margin-bottom: 10px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 14px;
            }}

            .upload-box button {{
                background: #0b5fa5;
                color: white;
                border: none;
                cursor: pointer;
                font-weight: bold;
            }}

            .upload-box button:hover {{
                background: #0d47a1;
            }}

            .item-card {{
                background: white;
                padding: 20px;
                margin-bottom: 20px;
                border-radius: 8px;
                box-shadow: 0 3px 8px rgba(0,0,0,0.15);
                display: flex;
                justify-content: space-between;
                gap: 20px;
                align-items: center;
            }}

            .item-left h3 {{
                margin: 0;
                color: #0b5fa5;
            }}

            .btn-download {{
                background: #4CAF50;
                color: white;
                padding: 10px 15px;
                border-radius: 4px;
                text-decoration: none;
                font-weight: bold;
                display: inline-block;
                cursor: pointer;
                transition: background 0.3s;
            }}

            .btn-download:hover {{
                background: #45a049;
            }}

            .back {{
                background: #0b5fa5;
                color: white;
                padding: 10px 18px;
                border-radius: 6px;
                text-decoration: none;
                font-weight: bold;
                display: inline-block;
            }}

            .back:hover {{
                background: #0d47a1;
            }}

            .message {{
                padding: 10px;
                margin-bottom: 10px;
                border-radius: 4px;
                text-align: center;
            }}

            .success {{
                background: #d4edda;
                color: #155724;
                border: 1px solid #c3e6cb;
            }}

            .error {{
                background: #f8d7da;
                color: #721c24;
                border: 1px solid #f5c6cb;
            }}

            @media (max-width: 768px) {{
                .container {{
                    flex-direction: column;
                }}

                .sidebar {{
                    width: 100%;
                }}
            }}
             /* ===== ANIMASI TEKS HEADER ===== */
            .header {{
              overflow: hidden;
        }}
            .header .running-text {{
            display: inline-block;
             white-space: nowrap;
             animation: slideText 25s linear infinite;
            }}

            @keyframes slideText {{
            0% {{
             transform: translateX(100%);
             }}
             100% {{
           transform: translateX(-100%);
             }}
    }}
        </style>
    </head>
    <body>

        <div class="header">
        <div class="running-text">
        <span>
        <h1> INVENTORI LAB LISTRIK JURUSAN TEKNIK MESIN POLITEKNIK NEGERI SUBANG</h1>
            </span>
            </div>
        </div>

        <div class="container">
            <div class="list">
                <h2>Daftar PDF</h2>
                {items_html if items_html.strip() else '<p>Belum ada file PDF</p>'}
            </div>

            <div class="sidebar">
                <div class="upload-box">
                    <h3>📤 Upload PDF</h3>
                    <form method="POST" enctype="multipart/form-data" action="/upload_pdf_b">
                        <input type="text" name="nama" placeholder="Nama File" required>
                        <input type="file" name="file" accept=".pdf" required>
                        <button type="submit">Upload</button>
                    </form>
                </div>
                <a href="/kembali-dashboard" class="back">⬅ Kembali</a>
            </div>
        </div>

    </body>
    </html>
    """

@app.route('/upload_pdf_b', methods=['POST'])
def upload_pdf_b():
    if session.get('role') != 'admin':
        return redirect(url_for('inventori_b'))

    if 'file' not in request.files:
        return redirect(url_for('inventori_b'))
    
    file = request.files['file']
    nama = request.form.get('nama', 'file')
    
    if file.filename == '' or not allowed_file(file.filename):
        return redirect(url_for('inventori_b'))
    
    filename = secure_filename(f"b_{nama}.pdf")
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    return redirect(url_for('inventori_b'))

# ================= INVENTORI A / UMUM =================
@app.route('/inventori/<nama>')
def inventori(nama):

    # Allow viewing inventory contents without login (read-only)
    items = Barang.query.filter_by(ruang=nama.upper()).all()
    return render_template('inventori.html', nama=nama.upper(), items=items)

# ================= AMBIL BARANG =================
@app.route('/ambil/<int:item_id>', methods=['POST'])
def ambil_barang(item_id):
    if 'role' not in session:
        return redirect(url_for('dashboard'))

    item = Barang.query.get_or_404(item_id)

    if item.jumlah <= 0:
        flash('❌ Stok tidak tersedia')
        return redirect(request.referrer)

    item.jumlah -= 1
    db.session.commit()
    
    log = LogTransaksi(
        pengguna=session.get('username'),
        tanggal=datetime.now().strftime('%Y-%m-%d'),
        jam=datetime.now().strftime('%H:%M:%S'),
        aksi='Ambil',
        jumlah=1,
        unit=item.unit,
        jumlah_saat_ini=item.jumlah
    )
    
    db.session.add(log)
    db.session.commit()

    # Notifikasi jika stok menjadi rendah setelah diambil
    if item.jumlah <= 3:
        flash(f'⚠️ Stok {item.nama} tinggal {item.jumlah} - Segera Restock!')
    else:
        flash(f'✓ Berhasil mengambil {item.nama}')

    return redirect(request.referrer)

# ================= TAMBAH JUMLAH (ADMIN) =================
@app.route('/tambah/<int:item_id>', methods=['POST'])
def tambah_jumlah_barang(item_id):
    if session.get('role') != 'admin':
        return redirect(request.referrer)

    item = Barang.query.get_or_404(item_id)

    if item.jumlah >= 10:
        flash('❌ Stok sudah maksimal (10)')
        return redirect(request.referrer)

    item.jumlah += 1
    db.session.commit()
    
    log = LogTransaksi(
        pengguna=session.get('username'),
        tanggal=datetime.now().strftime('%Y-%m-%d'),
        jam=datetime.now().strftime('%H:%M:%S'),
        aksi='Tambah',
        jumlah=1,
        unit=item.unit,
        jumlah_saat_ini=item.jumlah
    )
    
    db.session.add(log)
    db.session.commit()

    if item.jumlah >= 10:
        flash(f'✓ Stok {item.nama} sudah Penuh (10)')
    else:
        flash(f'✓ Berhasil menambah {item.nama} (sekarang {item.jumlah})')

    return redirect(request.referrer)

# ================= TAMBAH BARANG BARU (ADMIN) =================
@app.route('/tambah_barang', methods=['POST'])
def tambah_barang_baru():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))

    barang = Barang(
        nama=request.form['nama'],
        jumlah=int(request.form['jumlah']),
        unit=request.form['unit'],
        ruang=request.form['ruang']
    )

    db.session.add(barang)
    db.session.commit()

    flash('Barang berhasil ditambahkan')
    return redirect(request.referrer)

# ================= RUN =================
if __name__ == "__main__":
    app.run(debug=True)